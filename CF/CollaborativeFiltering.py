from math import fabs
import re
import pandas as pd
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.preprocessing import normalize
from scipy import sparse
from flask import jsonify
import json
import openpyxl
import pprint


r_cols = ['user_id', 'item_id', 'rating']
uuCF = 1
k = 2
dist_func = cosine_similarity
Y_data = pd.read_excel('./CF/data.xlsx', names=r_cols).values
Ybar_data = Y_data.copy()
# n_users = len(np.array(np.unique(Y_data[:, 0])))

# n_items = len(np.array(np.unique(Y_data[:, 1])))
list_users = np.array(np.unique(Y_data[:, 0]))
list_items = np.array(np.unique(Y_data[:, 1]))
n_users = len(list_users)
n_items = len(list_items)
mu = np.zeros((n_users,))
Ybar = sparse.coo_matrix((Ybar_data[:, 2],
                          (Ybar_data[:, 1], Ybar_data[:, 0])), (int(np.max(Y_data[:, 1])) + 1, int(np.max(Y_data[:, 0])) + 1)).tocsr()
S = cosine_similarity(Ybar.T, Ybar.T, dense_output=False)


def reload_data():
    Y_data = pd.read_excel('./CF/data.xlsx', names=r_cols).values
    Ybar_data = Y_data.copy()
    n_users = len(np.array(np.unique(Y_data[:, 0])))
    n_items = len(np.array(np.unique(Y_data[:, 1])))
    mu = np.zeros((n_users,))
    Ybar = sparse.coo_matrix((Ybar_data[:, 2],
                              (Ybar_data[:, 1], Ybar_data[:, 0])), (int(np.max(Y_data[:, 1])) + 1, int(np.max(Y_data[:, 0])) + 1)).tocsr()
    S = dist_func(Ybar.T, Ybar.T, dense_output=False)
    list_users = np.array(np.unique(Y_data[:, 0]))
    list_items = np.array(np.unique(Y_data[:, 1]))


def add(new_data):
    Y_data = np.concatenate((Y_data, new_data), axis=0)


def normalize_Y():
    users = Y_data[:, 0]
    for n in range(0, n_users - 1):
        ids = np.where(users == n)[0].astype(np.int32)
        item_ids = Y_data[ids, 1]
        ratings = Y_data[ids, 2]
        mu[n] = np.mean(ratings)
        Ybar_data[ids, 2] = ratings - mu[n]
    # Ybar = sparse.coo_matrix((Ybar_data[:, 2],
    #                           (Ybar_data[:, 1], Ybar_data[:, 0])), (n_items, n_users))
    # Ybar = Ybar.tocsr()


def similarity():
    S = dist_func(Ybar.T, Ybar.T, dense_output=False)


def fit():
    normalize_Y()
    similarity()


def pred(u, i, normalized=1):
    ids = np.where(Y_data[:, 1] == i)[0].astype(np.int32)
    users_rated_i = (Y_data[ids, 0]).astype(np.int32)
    sim = S[u, users_rated_i].data
    a = np.argsort(sim)[-k:]
    nearest_s = sim[a]
    r = Ybar[i, users_rated_i[a]]
    if normalized:
        return (r*nearest_s)[0]/np.abs(nearest_s).sum()
    return (r*nearest_s)[0]/np.abs(nearest_s).sum() + mu[n]


def recommend(u):
    ids = np.where(Y_data[:, 0] == u)[0]
    items_rated_by_u = Y_data[ids, 1].tolist()
    recommended_items = []
    for i in list_items:
        if i not in items_rated_by_u:
            rating = pred(u, i)
            if rating > 0:
                recommended_items.append(i)

    return recommended_items


def print_recommendation():
    reload_data()
    fit()
    list_users = np.array(np.unique(Y_data[:, 0]))
    print('Recommendation: ')
    for u in list_users:
        recommended_items = recommend(u)
        print('    for user ', u, ': ', recommended_items)


def get_recommendation():
    reload_data()
    fit()
    arr = []
    for u in list_users:
        recommended_items = dict()
        result = dict()
        result['user_id'] = u
        result['film_id_array'] = recommend(u)
        recommended_items[u] = result
        arr.append(recommended_items)
    return arr
    # return json.dumps([item.__dict__ for item in arr])


def get_recommendation_by_user(user_id):
    reload_data()
    fit()
    result = dict()
    for u in list_users:
        if (u == user_id):
            result['user_id'] = u
            result['film_id_array'] = recommend(u)
    return result


def update_data(user_id, item_id, rating):
    path = 'CF\data.xlsx'
    Y_data = pd.read_excel(path, names=r_cols).values
    max_row = len(Y_data)
    max_column = 3
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['Sheet1']
    sheet_obj = wb_obj.active
    isExisted = False
    for i in range(1, max_row + 1):
        for j in range(1, max_column + 1):
            USER_ID = sheet_obj.cell(row=i, column=1)
            ITEM_ID = sheet_obj.cell(row=i, column=2)
            if USER_ID == user_id and ITEM_ID == item_id:
                sheet.cell(row=i, column=3, value=rating)
                isExisted = True
                break
    if isExisted == False:
        sheet.cell(row=max_row + 1, column=1, value=user_id)
        sheet.cell(row=max_row+1, column=2, value=item_id)
        sheet.cell(row=max_row+1, column=3, value=rating)
    wb_obj.save(path)
    return "OK"


# print(get_recommendation_by_user(100000))
#update_data(100000, 3242424, 2)
# get_recommendation_by_user(7)
# get_recommendation()
# print_recommendation()
#print(np.array(np.unique(Y_data[:, 0])))

# list_users = np.array(np.unique(Y_data[:, 0]))
# for item in list_users:
#     print(item)


def int_from_object_id(object_id):
    firstObjectId = '5661728913124370191fa3f8'
    delta = int(object_id[0: 8], 16)-int(firstObjectId[0: 8], 16)
    res = str(delta) + str(int(object_id[18: 24], 16))
    return int(res)


# print(int_from_object_id('62488aa75a3bafebc42c30a7'))
